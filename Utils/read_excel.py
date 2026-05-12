from openpyxl import load_workbook
wb = load_workbook("utils/sample.xlsx")
sheet = wb.active
for row in sheet.iter_rows(values_only=True):
    print(row)